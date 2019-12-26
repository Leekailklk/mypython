#import docx
from docx import Document #导入库
from Get_ID_info import GetInformation
import openpyxl

path = "4.4学生补录信息表20190918(3)(1)(1).docx" #文件路径
wb = openpyxl.Workbook()
ws = wb.active

document = Document(path) #读入文件
tables = document.tables #获取文件中的表格集
table = tables[0]#获取文件中的第一个表格
for i in range(0,len(table.rows)):#从表格第一行开始循环读取表格数据
    result = []
    for j in range(0,6):
        result .append(table.cell(i, j).text.strip('\n').replace(' ', ''))
    # cell(i,0)表示第(i+1)行第1列数据，以此类推
        #print(i,j, table.cell(i, j).text.strip('\n').replace(' ', ''))
    if i==0:
        result.append('父亲年龄')
        result.append('母亲年龄')
    else:
        #print(result[3],result[5])
        result.append(GetInformation(result[3]).get_age())
        result.append(GetInformation(result[5]).get_age())
    ws.append(result)
wb.save('students.xlsx')