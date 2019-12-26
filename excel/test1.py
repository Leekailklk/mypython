import openpyxl

input_file_name = 'rb201911.xlsx'


def read_excel(input_file_name):
    """
    从xlsx文件中读取数据
    """
    workbook = openpyxl.load_workbook(input_file_name)
    print(workbook)
    # 可以使用workbook对象的sheetnames属性获取到excel文件中哪些表有数据
    print(workbook.sheetnames)
    table = workbook.active
    print(table)
    rows = table.max_row
    cols = table.max_column

    for row in range(rows):
        for col in range(cols):
            data = table.cell(row + 1, col + 1).value
            print(data, end=' ')


read_excel(input_file_name)